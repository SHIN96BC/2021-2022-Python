#from typing import Container
#from urllib.request import urlopen 
import openpyxl as opxl
import requests as rqs
from bs4 import BeautifulSoup as bfs
from selenium import webdriver 
import time
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QComboBox, QLineEdit, QTextBrowser, QPushButton
from PyQt5.QtCore import Qt


searchStart = False
searchEnd = False
inputSearchFlag = False
inputSaveFlag = False
indexSetFlag = False
webChange = ""
search = ""
fnameInput = ""
searchIndex = ""

# GUI
class ExcelGui(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
    def initUI(self): 
        self.cb = QComboBox(self)
        self.cb.addItem("")
        self.cb.addItem("naver shopping")
        self.cb.move(10, 10)
        self.cb.activated[str].connect(self.changeWeb)
        
        self.inputSearch = QLineEdit(self)
        self.inputSearch.move(140,10)
        self.inputSearch.setText("검색어를 입력해주세요")
        self.inputSearch.cursorPositionChanged.connect(self.inputSearchSet)
        
        self.indexNum = QLineEdit(self)
        self.indexNum.move(10,40)
        self.indexNum.setText("검색할 페이지 수")
        self.indexNum.cursorPositionChanged.connect(self.inputIndexSet)
        
        self.inputSave = QLineEdit(self)
        self.inputSave.move(10,360)
        self.inputSave.setText("저장할 파일 이름을 입력해주세요")
        self.inputSave.cursorPositionChanged.connect(self.inputSaveSet)
        self.inputSave.setEnabled(False)
        
        self.output = QTextBrowser(self)
        self.output.setAcceptRichText(False)
        self.output.move(10, 70)
        self.output.setFixedSize(280, 270)
        
        bSearch = QPushButton(self)
        bSearch.move(190, 38)
        bSearch.setFixedSize(100, 25)
        bSearch.setText("검색")
        bSearch.released.connect(self.searchOption)
        
        self.bSave = QPushButton(self)
        self.bSave.move(190,358)
        self.bSave.setFixedSize(100, 25)
        self.bSave.setText("저장")
        self.bSave.released.connect(self.saveOption)
        self.bSave.setEnabled(False)
        
        self.setWindowTitle("webSearch")
        self.move(300, 300)
        self.setFixedSize(300, 400)
        
        
        self.show()
    def inputSearchSet(self):
        global inputSearchFlag
        if inputSearchFlag== False:
            self.inputSearch.setText("")
            inputSearchFlag = True
        elif len(self.inputSearch.text()) == 0:
            self.inputSearch.setText("검색어를 입력해주세요")
            inputSearchFlag = False
            
    def inputSaveSet(self):
        global inputSaveFlag
        if inputSaveFlag== False:
            self.inputSave.setText("")
            inputSaveFlag = True
        elif len(self.inputSave.text()) == 0:
            self.inputSave.setText("저장할 파일 이름을 입력해주세요")
            inputSaveFlag = False
    
    def inputIndexSet(self):
        global indexSetFlag
        if indexSetFlag== False:
            self.indexNum.setText("")
            indexSetFlag = True
        elif len(self.indexNum.text()) == 0:
            self.indexNum.setText("검색할 페이지 수")
            indexSetFlag = False
       
    def changeWeb(self, text):
        global webChange
        webChange = text
        
    def searchSet(self):
        global search
        global inputSearchFlag
        if self.inputSearch.text() == "검색어를 입력해주세요" or len(self.inputSearch.text()) == 0:
            self.output.append("검색어를 입력하세요")
            inputSearchFlag = False
        else: 
            search = self.inputSearch.text()
            self.output.append("검색어: "+ search)
            self.inputSearch.setText("검색어를 입력해주세요")
            inputSearchFlag = False
            return True
        
    def saveSet(self):
        global fnameInput
        global inputSaveFlag
        if self.inputSave.text() == "저장할 파일 이름을 입력해주세요" or len(self.inputSave.text()) == 0:
            self.output.append("저장할 파일 이름을 입력하세요")
            inputSaveFlag = False
        else: 
            fnameInput = self.inputSave.text()
            self.inputSave.setText("저장할 파일 이름을 입력해주세요")
            self.output.append("파일 이름: "+ fnameInput)
            inputSaveFlag = False
            return True
    
    def indexSet(self):
        global searchIndex
        global indexSetFlag
        if self.indexNum.text() == "검색할 페이지 수" or len(self.indexNum.text()) == 0:
            self.output.append("검색할 페이지 수를 입력하세요")
            indexSetFlag = False
        else: 
            try:
                int(self.indexNum.text())
                searchIndex = self.indexNum.text()
                self.indexNum.setText("검색할 페이지 수")
                self.output.append("페이지 수: "+ searchIndex)
                indexSetFlag = False
                return True
            except ValueError:
                self.output.append("페이지수는 숫자만 입력하세요")
                self.indexNum.setText("검색할 페이지 수")
            
    def searchOption(self):
        global searchStart 
        global webChange
        searchSetF = self.searchSet()
        indexSetF = self.indexSet()
        if searchSetF and indexSetF:
            if webChange == "naver shopping":
                searchStart = True
                self.inputSave.setEnabled(False)
                self.bSave.setEnabled(False)
                self.naverShopping()
                self.output.append("검색완료 완료!!")
    
    def saveOption(self):
        global searchStart 
        saveSetF = self.saveSet()
        if saveSetF and searchStart == False:
            self.excelSave()
                   
    def naverShopping(self):
        global searchStart
        global searchIndex
        global search
        
        if searchStart:
            self.wb = opxl.Workbook()
            sheet = self.wb.active
            sheet.append(["물건이름", "물건종류", "가격"])
            option = webdriver.ChromeOptions()
            option.add_argument("window-size=1920,1080")
            driver = webdriver.Chrome("chromedriver.exe", options=option)
            driver.implicitly_wait(time_to_wait=5)

            for p in range(1,int(searchIndex)+1,1):
                url = "https://search.shopping.naver.com/search/all?frm=NVSCPRO&origQuery={s}&pagingIndex={p}&pagingSize=40&productSet=total&query={s}&sort=rel&timestamp=&viewType=list".format(s = search, p = p)
                last_scroll = driver.execute_script("return document.body.scrollHeight")
                driver.get(url)
                
                while True:    # 스크롤 자동으로 내려줘서 자바스크립트때문에 안나오던 div들을 가져옴
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(1)
                    new_height = driver.execute_script("return document.body.scrollHeight")
                    if new_height == last_scroll :
                        break
                    last_scroll = new_height
                
                html = driver.page_source
                soup = bfs(html, "html.parser")
                clips = soup.select("div.basicList_info_area__17Xyo")
                
                for cl in clips:
                    name = cl.select_one("div.basicList_title__3P9Q7").text.strip()   # 이름
                    price = cl.select_one("div.basicList_price_area__1UXXR").text.strip()  # 종류
                    Kinds = cl.select_one("div.basicList_depth__2QIie").text.strip()  # 가격
                    sheet.append([name, Kinds, price])
                    searchStart = False
                    self.inputSave.setEnabled(True)
                    self.bSave.setEnabled(True)

    def excelSave(self):
        global fnameInput
        
        fname = "{f}.xlsx".format(f = fnameInput)
        self.wb.save(fname)
        self.output.append("저장 완료!![ " + fname + " ]")
# 실행      
app = QApplication(sys.argv)
ex = ExcelGui()
sys.exit(app.exec_())

'''
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelGui()
    sys.exit(app.exec_())
'''

'''
# 내가 작업할 엑셀파일 생성
wb = opxl.Workbook()

# 작업할 엑셀파일 내 시트 활성화
sheet = wb.active

search = input("검색할 데이터를 입력해주세요: ")

sheet.append(["검색어명", "기사 제목", "기사 요약"])
#raw = rqs.get("https://tv.naver.com/r")
for p in range(1, 10+1, 1):
    raw = rqs.get("https://search.daum.net/search?w=news&DA=PGD&enc=utf8&cluster=y&cluster_page=1&q={s}&p={p}".format(s = search, p = p))
    html = bfs(raw.text, "html.parser")

Container = html.select("ul#clusterResultUL li")

for con in Container:
    #t = con.select_one("dt.title").text.strip()
    #c = con.select_one("dd.chn").text.strip()
    #h = con.select_one("span.hit").text.strip()
    #l = con.select_one("span.like").text.strip()
    #sheet.append([t,c,h,l])
    title = con.select_one("div.wrap_tit.mg_tit a").text.strip()
    content = con.select_one("p.f_eb.desc").text.strip()
    sheet.append([search, title, content])

# 시트 내 셀 선택 (A1이라는 셀에 1을 할당)
sheet['A1'] = 1

# 행번호, 열번호로 데이터 입력하기
#sheet.cell(row = 1, column = 1).value = 1

# 작업 마친 후 파일 저장
wb.save("파일명.xlsx")
'''



'''
for n in range(1,11,1):
    raw = rqs.get("https://search.shopping.naver.com/search/all?frm=NVSCPRO&origQuery={s}&pagingIndex="+str(n)+"&pagingSize=40&productSet=total&query={s}&sort=rel&timestamp=&viewType=list".format(s = search))
    html = bfs(raw.text, "html.parser")
    html.find_all("div")
    #clips = html.select("div.basicList_info_area__17Xyo")
    clips = html.select("div.basicList_info_area__17Xyo")
    print(clips)

    for cl in clips:
        #t = con.select_one("dt.title").text.strip()
        #c = con.select_one("dd.chn").text.strip()
        #h = con.select_one("span.hit").text.strip()
        #l = con.select_one("span.like").text.strip()
        #sheet.append([t,c,h,l])
        #name = cl.select_one("a.basicList_link_1MaTN")
        #price = cl.select_one("span.price_num__2WUXn")
        #Kinds = cl.select_one("a.basiclist_category_wVevj")
        name = cl.select_one("div.basicList_title__3P9Q7").text.strip()
        price = cl.select_one("div.basicList_price_area__1UXXR").text.strip()
        Kinds = cl.select_one("div.basicList_depth__2QIie").text.strip()
        sheet.append([name, Kinds, price])

wb.save("1.xlsx")
'''

#driver = webdriver.Chrome(execcutable_path="")